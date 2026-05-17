from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional
from uuid import UUID
from datetime import date
import csv
import io
import json
import re
import logging
from urllib.parse import urlparse

import httpx
from bs4 import BeautifulSoup

from sqlalchemy import select

from app.database import get_db
from app.models.user import User
from app.models.relevance import OrganizationProfile
from app.schemas.device import (
    DeviceCreate, DeviceUpdate, DeviceResponse,
    DeviceListResponse, DeviceSearchParams,
    BulkActionRequest, BulkActionResult,
)
from app.dependencies import get_current_user, get_optional_current_user, require_role
from app.services.device_service import DeviceService
from app.services.opportunity_relevance_service import OpportunityRelevanceService
from app.utils.text_utils import localize_investment_text, looks_english_text

logger = logging.getLogger(__name__)

# ─── Scraper helper ─────────────────────────────────────────────────────────

SKIP_TAGS = {"script", "style", "nav", "footer", "header", "aside",
             "noscript", "iframe", "form", "button", "svg", "img"}

ELIGIBILITY_KEYWORDS = [
    "éligib", "eligib", "bénéficiaire", "beneficiaire", "qui peut",
    "conditions", "critère", "critere", "exigence", "prérequis",
    "profil", "cible", "destiné", "destine", "pour qui",
]
AMOUNT_KEYWORDS = [
    "montant", "financement", "subvention", "aide", "dotation",
    "budget", "plafond", "jusqu'à", "jusqu'a", "entre ",
]
DATE_KEYWORDS = [
    "date", "clôture", "cloture", "délai", "delai", "dépôt",
    "dossier", "candidature", "ouverture", "fermeture",
]
PROCEDURE_KEYWORDS = [
    "comment postuler", "dossier", "démarche", "démarches",
    "étapes", "procedure", "procédure", "comment bénéficier",
    "comment candidater", "soumettre",
]


def _clean_text(text: str) -> str:
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


def _is_public_scrapable_url(url: str | None) -> bool:
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return False

    parsed = urlparse(url)
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()

    if host.startswith("api.") or ".api." in host:
        return False
    if path.startswith("/api") or "/api/" in path:
        return False
    if any(token in host for token in ("swagger", "openapi")):
        return False

    return True


def _find_public_url_in_payload(payload) -> str | None:
    if isinstance(payload, dict):
        for value in payload.values():
            candidate = _find_public_url_in_payload(value)
            if candidate:
                return candidate
        return None

    if isinstance(payload, list):
        for value in payload:
            candidate = _find_public_url_in_payload(value)
            if candidate:
                return candidate
        return None

    if isinstance(payload, str) and payload.startswith("http") and _is_public_scrapable_url(payload):
        return payload

    return None


def _resolve_scrapeable_source_url(device) -> str | None:
    if _is_public_scrapable_url(device.source_url):
        return device.source_url

    if not device.source_raw:
        return None

    try:
        payload = json.loads(device.source_raw)
    except Exception:
        return None

    return _find_public_url_in_payload(payload)


def _is_metadata_backed_device(device) -> bool:
    source_url = (device.source_url or "").lower()
    source_raw = device.source_raw or ""

    if "projects.worldbank.org" in source_url:
        return True

    return any(token in source_raw for token in ('"project_name"', '"closingdate"', '"countryshortname"', '"totalcommamt"'))


def _extract_text_block(el) -> str:
    text = el.get_text(" ", strip=True)
    if not text:
        return ""
    if el.name == "li":
        return f"- {text}"
    return text


def _pick_main_content(soup: BeautifulSoup):
    selectors = [
        "main article",
        "main .column_content",
        "main .node__content",
        "main .content-wrapper",
        "main .article-content",
        "main .entry-content",
        "main .rich-text",
        "main .wysiwyg",
        "main .prose",
        "main .page-body",
        "main",
    ]

    best = None
    best_len = 0
    for selector in selectors:
        for node in soup.select(selector):
            text_len = len(node.get_text(" ", strip=True))
            if text_len > best_len:
                best = node
                best_len = text_len

    return best or soup.find("main") or soup.body or soup


def _extract_sections(soup: BeautifulSoup, base_url: str) -> dict:
    """Extrait intelligemment les sections clés d'une page HTML."""
    # Supprimer les balises inutiles
    for tag in soup.find_all(SKIP_TAGS):
        tag.decompose()
    for tag in soup.find_all(attrs={"class": re.compile(
            r"(cookie|banner|popup|modal|menu|breadcrumb|sidebar|widget|social|share|comment)",
            re.I)}):
        tag.decompose()

    # Trouver le contenu principal
    main = _pick_main_content(soup)

    # Extraire tous les blocs de texte structurés (h2/h3 + texte suivant)
    sections: dict[str, list[str]] = {}
    current_title = "description"
    current_parts: list[str] = []

    for el in main.find_all(["h1", "h2", "h3", "h4", "p", "li"]):
        if el.name in ("h1", "h2", "h3", "h4"):
            if current_parts:
                sections.setdefault(current_title, []).extend(current_parts)
            current_title = el.get_text(strip=True).lower()
            current_parts = []
        else:
            txt = _extract_text_block(el)
            if len(txt) > 20:
                current_parts.append(txt)

    if current_parts:
        sections.setdefault(current_title, []).extend(current_parts)

    if not sections:
        fallback_parts = []
        for el in main.find_all(["p", "li"]):
            txt = _extract_text_block(el)
            if len(txt) > 20:
                fallback_parts.append(txt)
        if fallback_parts:
            sections["description"] = fallback_parts

    # Classifier chaque section
    eligibility_parts: list[str] = []
    amount_parts: list[str] = []
    date_parts: list[str] = []
    procedure_parts: list[str] = []
    description_parts: list[str] = []

    for title, parts in sections.items():
        text_block = "\n".join(parts)
        if any(kw in title for kw in ELIGIBILITY_KEYWORDS):
            eligibility_parts.extend(parts)
        elif any(kw in title for kw in AMOUNT_KEYWORDS):
            amount_parts.extend(parts)
        elif any(kw in title for kw in DATE_KEYWORDS):
            date_parts.extend(parts)
        elif any(kw in title for kw in PROCEDURE_KEYWORDS):
            procedure_parts.extend(parts)
        else:
            description_parts.extend(parts)

    if not description_parts:
        fallback_parts = []
        for el in main.find_all(["p", "li"]):
            txt = _extract_text_block(el)
            if len(txt) > 30:
                fallback_parts.append(txt)
        description_parts.extend(fallback_parts[:16])

    # Construire la description complète formatée
    full_desc_parts = []
    if description_parts:
        full_desc_parts.append("## Présentation\n" + "\n".join(description_parts[:8]))
    if amount_parts:
        full_desc_parts.append("## Montants & Financement\n" + "\n".join(amount_parts[:6]))
    if eligibility_parts:
        full_desc_parts.append("## Critères d'éligibilité\n" + "\n".join(eligibility_parts[:10]))
    if date_parts:
        full_desc_parts.append("## Dates & Délais\n" + "\n".join(date_parts[:6]))
    if procedure_parts:
        full_desc_parts.append("## Comment postuler\n" + "\n".join(procedure_parts[:8]))

    full_description = _clean_text("\n\n".join(full_desc_parts))[:8000]
    eligibility = _clean_text("\n".join(eligibility_parts))[:3000] if eligibility_parts else None

    return {
        "full_description": full_description or None,
        "eligibility_criteria": eligibility,
        "eligible_expenses": _clean_text("\n".join(amount_parts))[:2000] if amount_parts else None,
    }


async def _scrape_device_url(url: str) -> dict:
    """Scrape une URL et retourne les sections extraites."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    async with httpx.AsyncClient(
        timeout=20, follow_redirects=True,
        headers=headers, verify=False
    ) as client:
        resp = await client.get(url)
        resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "lxml")
    return _extract_sections(soup, url)

router = APIRouter(prefix="/api/v1/devices", tags=["devices"])


def _can_use_admin_catalog(current_user: User | None) -> bool:
    if not current_user:
        return False
    return current_user.role == "admin" or getattr(current_user, "platform_role", "member") == "super_admin"


def _is_standard_user(current_user: User | None) -> bool:
    if not current_user:
        return False
    return not _can_use_admin_catalog(current_user)


def _remove_unqualified_public_types(device_types: Optional[List[str]]) -> Optional[List[str]]:
    qualified_public_types = [
        "subvention",
        "pret",
        "avance_remboursable",
        "garantie",
        "credit_impot",
        "exoneration",
        "aap",
        "appel_a_projets",
        "ami",
        "accompagnement",
        "concours",
        "investissement",
    ]
    if not device_types:
        return qualified_public_types
    cleaned = [item for item in device_types if item not in {"autre", "institutional_project"}]
    return cleaned or ["__no_type_match__"]


PROFILE_BENEFICIARY_MAP = {
    "entrepreneur": [
        "entreprise",
        "pme",
        "tpe",
        "mpme",
        "startup",
        "entrepreneur",
        "porteur projet",
        "porteur de projet",
        "jeune entrepreneur",
        "femme entrepreneure",
        "cooperative",
        "entreprise sociale",
        "exploitant agricole",
        "structure_accompagnement",
    ],
    "association": [
        "association",
        "ong",
        "osc",
        "organisation communautaire",
        "organisation locale",
        "cooperative",
        "mutuelle",
        "entreprise sociale",
    ],
    "collectivite": [
        "collectivite",
        "collectivite territoriale",
        "mairie",
        "commune",
        "region",
        "institution publique",
        "etat",
        "ministere",
        "acteur territorial",
    ],
    "consultant": [
        "entreprise",
        "pme",
        "tpe",
        "mpme",
        "startup",
        "entrepreneur",
        "association",
        "ong",
        "collectivite",
        "institution publique",
        "cooperative",
    ],
}


def _profile_beneficiaries(profile: OrganizationProfile | None) -> Optional[List[str]]:
    if not profile or not profile.organization_type:
        return None
    return PROFILE_BENEFICIARY_MAP.get(str(profile.organization_type).lower())


def resolve_onboarding_result_filters(
    total: int,
    institutional_total: int,
    device_types: Optional[List[str]],
    beneficiaries: Optional[List[str]] = None,
) -> tuple[int, Optional[List[str]]]:
    """
    Aligne le nombre affiche pendant l'onboarding avec la page ouverte ensuite.

    L'onboarding utilisateur ne doit pas compter les signaux institutionnels
    comme des opportunites actionnables. Ils restent exploitables en admin.
    """
    return total, device_types


@router.get("/", response_model=DeviceListResponse)
async def list_devices(
    q: Optional[str] = Query(None, description="Recherche plein texte"),
    countries: Optional[List[str]] = Query(None),
    device_types: Optional[List[str]] = Query(None),
    sectors: Optional[List[str]] = Query(None),
    beneficiaries: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    validation_status: Optional[str] = Query(None),
    closing_soon_days: Optional[int] = Query(None),
    has_close_date: Optional[bool] = Query(None),
    actionable_now: Optional[bool] = Query(None),
    close_date_before: Optional[date] = Query(None),
    close_date_after: Optional[date] = Query(None),
    min_ai_readiness: Optional[int] = Query(None, ge=0, le=100),
    ai_readiness_labels: Optional[List[str]] = Query(None),
    include_all_statuses: bool = Query(False),
    include_rejected: bool = Query(False),
    include_low_quality: bool = Query(False),
    sort_by: str = Query("updated_at"),
    sort_desc: bool = Query(True),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User | None = Depends(get_optional_current_user),
):
    admin_catalog_requested = include_all_statuses or include_rejected or include_low_quality
    if admin_catalog_requested and not _can_use_admin_catalog(current_user):
        raise HTTPException(status_code=403, detail="Le catalogue complet est reserve aux administrateurs.")

    # ── Filtre profil strict pour les utilisateurs normaux ─────────────────
    # Les admins voient tout (filtre pays manuel respecté).
    # Les utilisateurs normaux voient UNIQUEMENT les dispositifs de leurs pays
    # de profil — le filtre pays manuel est ignoré pour eux.
    # Si le profil n'existe pas ou n'a pas de pays → 0 résultats.
    effective_countries = countries  # utilisé tel quel pour les admins

    effective_beneficiaries = beneficiaries

    if _is_standard_user(current_user):
        # Pour les utilisateurs normaux : toujours appliquer le profil,
        # ignorer tout filtre pays manuel envoyé par le client.
        try:
            relevance_svc = OpportunityRelevanceService(db)
            org_id = await relevance_svc.get_current_organization_id(current_user)
            if org_id:
                prof_row = await db.execute(
                    select(OrganizationProfile).where(OrganizationProfile.organization_id == org_id)
                )
                profile = prof_row.scalar_one_or_none()
                if profile and profile.countries:
                    effective_countries = profile.countries
                    profile_beneficiaries = _profile_beneficiaries(profile)
                    if profile_beneficiaries:
                        if beneficiaries:
                            overlap = sorted(set(beneficiaries) & set(profile_beneficiaries))
                            effective_beneficiaries = overlap or ["__no_beneficiary_match__"]
                        else:
                            effective_beneficiaries = profile_beneficiaries
                else:
                    # Profil sans pays → 0 résultats (onboarding incomplet)
                    effective_countries = ["__no_country_match__"]
            else:
                # Pas d'organisation → 0 résultats
                effective_countries = ["__no_country_match__"]
        except Exception:
            # Erreur inattendue → ne pas exposer tous les dispositifs
            effective_countries = ["__no_country_match__"]

    effective_device_types = _remove_unqualified_public_types(device_types) if _is_standard_user(current_user) else device_types

    params = DeviceSearchParams(
        q=q, countries=effective_countries, device_types=effective_device_types,
        sectors=sectors, beneficiaries=effective_beneficiaries, status=status,
        validation_status=validation_status,
        closing_soon_days=closing_soon_days,
        has_close_date=has_close_date,
        actionable_now=actionable_now,
        close_date_before=close_date_before,
        close_date_after=close_date_after,
        min_ai_readiness=min_ai_readiness,
        ai_readiness_labels=ai_readiness_labels,
        include_all_statuses=include_all_statuses,
        include_rejected=include_rejected,
        include_low_quality=include_low_quality,
        sort_by=sort_by,
        sort_desc=sort_desc, page=page, page_size=page_size,
    )
    service = DeviceService(db)
    result = await service.search(params)
    if current_user:
        await OpportunityRelevanceService(db).attach_runtime_relevance(result["items"], user=current_user)
    return result


@router.get("/stats")
async def device_stats(db: AsyncSession = Depends(get_db)):
    return await DeviceService(db).get_stats()


@router.get("/onboarding-preview")
async def onboarding_preview(
    countries: Optional[List[str]] = Query(None),
    device_types: Optional[List[str]] = Query(None),
    sectors: Optional[List[str]] = Query(None),
    beneficiaries: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    actionable_now: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User | None = Depends(get_optional_current_user),
):
    """
    Calcule un aperçu réel pendant l'onboarding.

    Contrairement à /devices, cet endpoint n'applique pas le profil utilisateur
    déjà enregistré, car celui-ci n'existe pas encore forcément au moment de
    l'inscription. Il utilise uniquement les critères que l'utilisateur vient
    de sélectionner dans l'écran d'onboarding.
    """
    if not countries:
        return {
            "total": 0,
            "subventions": 0,
            "investisseurs": 0,
            "urgentes": 0,
            "has_real_count": True,
        }

    service = DeviceService(db)

    base_params = {
        "countries": countries,
        "device_types": device_types,
        "sectors": sectors,
        "beneficiaries": beneficiaries,
        "status": status,
        "actionable_now": actionable_now,
        "sort_by": "relevance",
        "page": 1,
        "page_size": 1,
    }

    total_result = await service.search(DeviceSearchParams(**base_params))
    subvention_types = ["subvention", "aap", "concours", "pret", "accompagnement", "garantie"]
    investor_types = ["investissement"]

    requested_types = set(device_types or [])
    count_public = not requested_types or bool(requested_types & set(subvention_types))
    count_private = not requested_types or bool(requested_types & set(investor_types))

    subventions = 0
    investisseurs = 0
    if count_public:
        public_types = sorted((requested_types & set(subvention_types)) or set(subvention_types))
        public_result = await service.search(DeviceSearchParams(**{**base_params, "device_types": public_types}))
        subventions = public_result["total"]
    if count_private:
        private_types = sorted((requested_types & set(investor_types)) or set(investor_types))
        private_result = await service.search(DeviceSearchParams(**{**base_params, "device_types": private_types}))
        investisseurs = private_result["total"]

    institutional_result = await service.search(
        DeviceSearchParams(**{**base_params, "device_types": ["institutional_project"]})
    )
    display_total, result_device_types = resolve_onboarding_result_filters(
        total_result["total"],
        institutional_result["total"],
        device_types,
        beneficiaries,
    )

    urgent_result = await service.search(
        DeviceSearchParams(
            **{
                **base_params,
                "status": ["open"],
                "closing_soon_days": 30,
            }
        )
    )

    return {
        "total": total_result["total"],
        "display_total": display_total,
        "subventions": subventions,
        "investisseurs": investisseurs,
        "institutionnels": institutional_result["total"],
        "urgentes": urgent_result["total"],
        "result_device_types": result_device_types,
        "result_status": status,
        "result_actionable_now": actionable_now,
        "has_real_count": True,
    }


EXPORT_FIELDS = [
    "titre", "organisme", "pays", "type_aide", "statut",
    "montant_min", "montant_max", "devise", "taux_financement",
    "date_ouverture", "date_cloture", "secteurs", "beneficiaires",
    "description_courte", "url_source", "score_fiabilite", "score_completude",
]

EXPORT_HEADERS_FR = {
    "titre": "Titre",
    "organisme": "Organisme",
    "pays": "Pays",
    "type_aide": "Type d'aide",
    "statut": "Statut",
    "montant_min": "Montant min (€)",
    "montant_max": "Montant max (€)",
    "devise": "Devise",
    "taux_financement": "Taux (%)",
    "date_ouverture": "Date d'ouverture",
    "date_cloture": "Date de clôture",
    "secteurs": "Secteurs",
    "beneficiaires": "Bénéficiaires",
    "description_courte": "Description",
    "url_source": "URL source",
    "score_fiabilite": "Fiabilité (%)",
    "score_completude": "Complétude (%)",
}


def _device_to_row(device) -> dict:
    """Sérialise un Device en dict prêt pour l'export."""
    return {
        "titre":             device.title or "",
        "organisme":         device.organism or "",
        "pays":              device.country or "",
        "type_aide":         device.device_type or "",
        "statut":            device.status or "",
        "montant_min":       float(device.amount_min) if device.amount_min is not None else "",
        "montant_max":       float(device.amount_max) if device.amount_max is not None else "",
        "devise":            device.currency or "EUR",
        "taux_financement":  device.funding_rate if device.funding_rate else "",
        "date_ouverture":    device.open_date.isoformat() if device.open_date else "",
        "date_cloture":      device.close_date.isoformat() if device.close_date else "",
        "secteurs":          ", ".join(device.sectors or []),
        "beneficiaires":     ", ".join(device.beneficiaries or []),
        "description_courte": (device.short_description or "")[:300],
        "url_source":        device.source_url or "",
        "score_fiabilite":   device.confidence_score or 0,
        "score_completude":  device.completeness_score or 0,
    }


@router.get("/export/csv")
async def export_csv(
    q: Optional[str] = Query(None),
    countries: Optional[List[str]] = Query(None),
    device_types: Optional[List[str]] = Query(None),
    sectors: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    ai_readiness_labels: Optional[List[str]] = Query(None),
    min_ai_readiness: Optional[int] = Query(None, ge=0, le=100),
    closing_soon_days: Optional[int] = Query(None),
    has_close_date: Optional[bool] = Query(None),
    actionable_now: Optional[bool] = Query(None),
    include_all_statuses: bool = Query(False),
    include_rejected: bool = Query(False),
    include_low_quality: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User | None = Depends(get_optional_current_user),
):
    """Export CSV streamé — BOM UTF-8 pour compatibilité Excel, max 5 000 lignes."""
    params = DeviceSearchParams(
        q=q, countries=countries, device_types=device_types,
        sectors=sectors, status=status, closing_soon_days=closing_soon_days,
        has_close_date=has_close_date,
        actionable_now=actionable_now,
        ai_readiness_labels=ai_readiness_labels,
        min_ai_readiness=min_ai_readiness,
        include_all_statuses=include_all_statuses,
        include_rejected=include_rejected,
        include_low_quality=include_low_quality,
    )
    admin_catalog_requested = include_all_statuses or include_rejected or include_low_quality
    if admin_catalog_requested and not _can_use_admin_catalog(current_user):
        raise HTTPException(status_code=403, detail="Le catalogue complet est reserve aux administrateurs.")

    service = DeviceService(db)

    async def generate_csv():
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
        # En-tête en français avec BOM
        buf.write("\ufeff")
        writer.writerow({k: EXPORT_HEADERS_FR.get(k, k) for k in EXPORT_FIELDS})
        yield buf.getvalue().encode("utf-8")

        count = 0
        async for device in service.stream_for_export(params, limit=5000):
            buf = io.StringIO()
            writer = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
            writer.writerow(_device_to_row(device))
            yield buf.getvalue().encode("utf-8")
            count += 1

        logger.info(f"[Export CSV] {count} dispositifs (q={q!r})")

    return StreamingResponse(
        generate_csv(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="kafundo-export.csv"',
            "Cache-Control": "no-store",
        },
    )


@router.get("/export/excel")
async def export_excel(
    q: Optional[str] = Query(None),
    countries: Optional[List[str]] = Query(None),
    device_types: Optional[List[str]] = Query(None),
    sectors: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    ai_readiness_labels: Optional[List[str]] = Query(None),
    min_ai_readiness: Optional[int] = Query(None, ge=0, le=100),
    closing_soon_days: Optional[int] = Query(None),
    has_close_date: Optional[bool] = Query(None),
    actionable_now: Optional[bool] = Query(None),
    include_all_statuses: bool = Query(False),
    include_rejected: bool = Query(False),
    include_low_quality: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User | None = Depends(get_optional_current_user),
):
    """Export Excel (.xlsx) avec formatage — max 5 000 lignes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé sur le serveur.")

    params = DeviceSearchParams(
        q=q, countries=countries, device_types=device_types,
        sectors=sectors, status=status, closing_soon_days=closing_soon_days,
        has_close_date=has_close_date,
        actionable_now=actionable_now,
        ai_readiness_labels=ai_readiness_labels,
        min_ai_readiness=min_ai_readiness,
        include_all_statuses=include_all_statuses,
        include_rejected=include_rejected,
        include_low_quality=include_low_quality,
    )
    admin_catalog_requested = include_all_statuses or include_rejected or include_low_quality
    if admin_catalog_requested and not _can_use_admin_catalog(current_user):
        raise HTTPException(status_code=403, detail="Le catalogue complet est reserve aux administrateurs.")

    service = DeviceService(db)

    # ── Collecter les données ─────────────────────────────────────────────
    rows = []
    async for device in service.stream_for_export(params, limit=5000):
        rows.append(_device_to_row(device))

    # ── Construire le classeur ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Kafundo Export"

    # Styles
    header_fill   = PatternFill("solid", fgColor="1D4ED8")   # bleu primary
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    subrow_fill   = PatternFill("solid", fgColor="EFF6FF")   # bleu très clair (lignes paires)
    border_side   = Side(style="thin", color="D1D5DB")
    cell_border   = Border(bottom=border_side)
    url_font      = Font(color="2563EB", underline="single", size=9)
    default_font  = Font(size=9)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # En-tête
    header_labels = [EXPORT_HEADERS_FR.get(f, f) for f in EXPORT_FIELDS]
    ws.append(header_labels)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_align

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Données
    for row_idx, row_data in enumerate(rows, start=2):
        values = [row_data.get(f, "") for f in EXPORT_FIELDS]
        ws.append(values)
        is_even = (row_idx % 2 == 0)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            field = EXPORT_FIELDS[col_idx - 1]
            if is_even:
                cell.fill = subrow_fill
            cell.border = cell_border
            if field == "url_source" and cell.value:
                cell.font = url_font
            else:
                cell.font = default_font
            cell.alignment = left_align

    # Largeurs de colonnes automatiques
    COL_WIDTHS = {
        "titre": 45, "organisme": 28, "pays": 16, "type_aide": 20,
        "statut": 12, "montant_min": 15, "montant_max": 15, "devise": 8,
        "taux_financement": 10, "date_ouverture": 14, "date_cloture": 14,
        "secteurs": 22, "beneficiaires": 22, "description_courte": 50,
        "url_source": 40, "score_fiabilite": 12, "score_completude": 12,
    }
    for col_idx, field in enumerate(EXPORT_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(field, 15)

    # Hauteur de ligne uniforme
    for row_idx in range(1, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 18

    # Filtre automatique sur l'en-tête
    ws.auto_filter.ref = ws.dimensions

    # Onglet récapitulatif
    ws_info = wb.create_sheet("Infos export")
    from datetime import datetime
    ws_info.append(["Export Kafundo"])
    ws_info.append(["Date", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws_info.append(["Filtres", f"q={q or ''} | pays={','.join(countries or [])} | types={','.join(device_types or [])}"])
    ws_info.append(["Nombre de lignes", len(rows)])
    ws_info["A1"].font = Font(bold=True, size=11)

    # ── Sérialiser et retourner ───────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info(f"[Export Excel] {len(rows)} dispositifs (q={q!r})")

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="kafundo-export.xlsx"',
            "Cache-Control": "no-store",
        },
    )


@router.post("/bulk", response_model=BulkActionResult)
async def bulk_action(
    body: BulkActionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    """
    Applique une action sur une sélection de dispositifs.
    - validate / reject / tag → admin ou editor
    - delete → admin uniquement
    """
    if not body.ids:
        raise HTTPException(status_code=422, detail="Aucun dispositif sélectionné.")
    if len(body.ids) > 200:
        raise HTTPException(status_code=422, detail="Maximum 200 dispositifs par action groupée.")

    if body.action == "delete" and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="La suppression groupée est réservée aux administrateurs.")

    if body.action == "tag" and not body.tags:
        raise HTTPException(status_code=422, detail="Au moins un tag est requis pour l'action 'tag'.")

    service = DeviceService(db)

    if body.action == "validate":
        return await service.bulk_validate(body.ids, current_user.id)
    elif body.action == "reject":
        return await service.bulk_reject(body.ids, current_user.id)
    elif body.action == "delete":
        return await service.bulk_delete(body.ids)
    elif body.action == "tag":
        return await service.bulk_tag(body.ids, body.tags, current_user.id)


@router.get("/{device_id}", response_model=DeviceResponse)
async def get_device(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User | None = Depends(get_optional_current_user),
):
    device = await DeviceService(db).get_by_id(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")
    if current_user:
        await OpportunityRelevanceService(db).attach_runtime_relevance([device], user=current_user)
    return device


@router.get("/{device_id}/history")
async def get_device_history(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    return await DeviceService(db).get_history(device_id)


@router.post("/", response_model=DeviceResponse, status_code=201)
async def create_device(
    data: DeviceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    return await DeviceService(db).create(data, created_by=str(current_user.id))


@router.put("/{device_id}", response_model=DeviceResponse)
async def update_device(
    device_id: UUID,
    data: DeviceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    device = await DeviceService(db).update(device_id, data, updated_by=str(current_user.id))
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")
    return device


@router.post("/{device_id}/validate", response_model=DeviceResponse)
async def validate_device(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    device = await DeviceService(db).validate(device_id, current_user.id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")
    return device


@router.post("/{device_id}/reject", response_model=DeviceResponse)
async def reject_device(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    device = await DeviceService(db).reject(device_id, current_user.id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")
    return device


@router.post("/{device_id}/scrape", response_model=DeviceResponse)
async def scrape_device_details(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    """
    Scrape l'URL officielle du dispositif et enrichit la fiche :
    description complète, critères d'éligibilité, dépenses éligibles.
    """
    service = DeviceService(db)
    device = await service.get_by_id(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")

    source_url = _resolve_scrapeable_source_url(device)
    if not source_url or source_url.startswith("https://www.google.com/search"):
        raise HTTPException(
            status_code=422,
            detail="Pas de page publique enrichissable disponible pour ce dispositif."
        )

    try:
        logger.info(f"[Scrape] {device.title} → {source_url}")
        scraped = await _scrape_device_url(source_url)
    except httpx.HTTPStatusError as e:
        if _is_metadata_backed_device(device) and e.response.status_code in {401, 403}:
            logger.info(f"[Scrape] Fallback metadata-only for '{device.title}'")
            return device
        if e.response.status_code == 401:
            raise HTTPException(
                status_code=422,
                detail="La fiche pointe vers une URL protégée ou non publique. L'enrichissement manuel nécessite une page source accessible publiquement."
            )
        if e.response.status_code == 404:
            raise HTTPException(
                status_code=422,
                detail="La page source officielle n'est plus disponible (404). "
                       "Le dispositif peut être archivé, expiré ou l'URL enregistrée n'est plus valide."
            )
        raise HTTPException(
            status_code=502,
            detail=f"Impossible d'accéder à l'URL ({e.response.status_code}). "
                   "Le site peut bloquer les requêtes automatiques."
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Erreur lors du scraping : {str(e)}")

    if not scraped.get("full_description") and not scraped.get("eligibility_criteria") and not scraped.get("eligible_expenses"):
        if _is_metadata_backed_device(device):
            logger.info(f"[Scrape] Aucun gain HTML pour '{device.title}', conservation de la fiche metadata")
            return device
        raise HTTPException(
            status_code=422,
            detail="Aucun contenu éditorial exploitable trouvé sur cette page. "
                   "Le site peut utiliser du JavaScript dynamique ou une structure HTML trop pauvre."
        )

    if device.device_type == "investissement":
        for field in ("full_description", "eligibility_criteria", "eligible_expenses"):
            value = scraped.get(field)
            if value and looks_english_text(value):
                scraped[field] = localize_investment_text(value)

    # Ne pas écraser des données déjà renseignées manuellement
    update_data = DeviceUpdate(
        full_description=(
            scraped.get("full_description")
            if scraped.get("full_description") and not device.full_description
            else device.full_description
        ),
        eligibility_criteria=(
            scraped.get("eligibility_criteria")
            if scraped.get("eligibility_criteria") and not device.eligibility_criteria
            else device.eligibility_criteria
        ),
        eligible_expenses=(
            scraped.get("eligible_expenses")
            if scraped.get("eligible_expenses") and not device.eligible_expenses
            else device.eligible_expenses
        ),
    )

    updated = await service.update(device_id, update_data, updated_by="scraper")
    logger.info(f"[Scrape] Fiche '{device.title}' enrichie avec succès")
    return updated


@router.post("/{device_id}/analyze", response_model=DeviceResponse)
async def analyze_device(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Génère une analyse décisionnelle IA pour aider l'utilisateur à décider
    vite : go/no-go, pourquoi intéressant, pourquoi prudent, action conseillée.
    """
    service = DeviceService(db)
    device = await service.get_by_id(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")

    # Construire le contexte du dispositif pour l'IA
    amount_info = ""
    if device.amount_min or device.amount_max:
        currency = device.currency or "EUR"
        if device.amount_min and device.amount_max:
            amount_info = f"entre {int(device.amount_min):,} et {int(device.amount_max):,} {currency}"
        elif device.amount_max:
            amount_info = f"jusqu'à {int(device.amount_max):,} {currency}"
        elif device.amount_min:
            amount_info = f"à partir de {int(device.amount_min):,} {currency}"
    elif device.funding_rate:
        amount_info = f"taux de financement : {device.funding_rate}%"

    deadline_info = ""
    if device.close_date:
        from datetime import date as date_class
        days_left = (device.close_date - date_class.today()).days
        if days_left < 0:
            deadline_info = f"Date de clôture dépassée ({device.close_date})"
        elif days_left <= 7:
            deadline_info = f"⚠️ URGENT — clôture dans {days_left} jour(s) ({device.close_date})"
        elif days_left <= 30:
            deadline_info = f"Clôture proche — dans {days_left} jours ({device.close_date})"
        else:
            deadline_info = f"Clôture le {device.close_date} ({days_left} jours restants)"
    elif device.is_recurring:
        deadline_info = "Dispositif récurrent / permanent"

    context_parts = [
        f"Titre : {device.title}",
        f"Organisme : {device.organism}",
        f"Pays : {device.country}" + (f", {device.region}" if device.region else ""),
        f"Type d'aide : {device.device_type}",
        f"Secteurs : {', '.join(device.sectors or []) or 'Non précisé'}",
        f"Bénéficiaires : {', '.join(device.beneficiaries or []) or 'Non précisé'}",
        f"Montant : {amount_info or 'Non précisé'}",
        f"Délai : {deadline_info or 'Non précisé'}",
        f"Statut : {device.status}",
    ]
    if device.short_description:
        context_parts.append(f"Description : {device.short_description[:500]}")
    if device.eligibility_criteria:
        context_parts.append(f"Critères d'éligibilité : {device.eligibility_criteria[:400]}")
    if device.specific_conditions:
        context_parts.append(f"Conditions spécifiques : {device.specific_conditions[:300]}")
    if device.required_documents:
        context_parts.append(f"Documents requis : {device.required_documents[:300]}")

    device_context = "\n".join(context_parts)

    prompt = f"""Tu es un expert en financement de projets africains et européens. Analyse ce dispositif de financement et génère une fiche décisionnelle structurée en JSON.

Dispositif à analyser :
{device_context}

Génère une analyse JSON avec exactement ces champs :
{{
  "go_no_go": "go" | "no_go" | "a_verifier",
  "recommended_priority": "haute" | "moyenne" | "faible",
  "why_interesting": "2-3 phrases sur ce qui est vraiment intéressant dans ce dispositif",
  "why_cautious": "2-3 phrases sur les points de vigilance, les risques ou les contraintes",
  "points_to_confirm": "2-3 points concrets à vérifier avant de candidater",
  "recommended_action": "1 phrase d'action concrète et directe pour l'utilisateur",
  "urgency_level": "critique" | "haute" | "moyenne" | "faible",
  "difficulty_level": "faible" | "moyenne" | "haute",
  "effort_level": "faible" | "moyenne" | "haute",
  "eligibility_score": <entier 0-100>,
  "strategic_interest": <entier 0-100>
}}

Règles :
- go_no_go = "go" si l'opportunité est clairement intéressante et accessible
- go_no_go = "no_go" si les critères sont trop restrictifs ou le dispositif peu adapté
- go_no_go = "a_verifier" si des informations manquent ou les critères sont flous
- urgency_level = "critique" si deadline < 7 jours, "haute" si < 30 jours
- difficulty_level = complexité administrative (documents, procédures)
- effort_level = temps et ressources nécessaires pour candidater
- eligibility_score = probabilité générale d'éligibilité (0=jamais éligible, 100=certainement éligible)
- strategic_interest = valeur stratégique du financement (montant, portée, récurrence)
- Sois concis, direct et actionnable. Pas de phrases vagues.
- Réponds UNIQUEMENT avec le JSON, sans texte autour."""

    try:
        import anthropic as _anthropic
        from app.config import settings

        api_key = getattr(settings, "ANTHROPIC_API_KEY", None)
        if not api_key:
            raise HTTPException(status_code=503, detail="API IA non configurée sur ce serveur.")

        client = _anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )

        raw_text = message.content[0].text.strip()
        # Nettoyer le JSON si besoin (parfois l'IA entoure de ```json```)
        if raw_text.startswith("```"):
            raw_text = raw_text.split("```")[1]
            if raw_text.startswith("json"):
                raw_text = raw_text[4:]
        raw_text = raw_text.strip()

        analysis = json.loads(raw_text)

        # Valider et normaliser les champs critiques
        valid_go = {"go", "no_go", "a_verifier"}
        valid_priority = {"haute", "moyenne", "faible"}
        valid_urgency = {"critique", "haute", "moyenne", "faible"}
        valid_level = {"faible", "moyenne", "haute"}

        if analysis.get("go_no_go") not in valid_go:
            analysis["go_no_go"] = "a_verifier"
        if analysis.get("recommended_priority") not in valid_priority:
            analysis["recommended_priority"] = "moyenne"
        if analysis.get("urgency_level") not in valid_urgency:
            analysis["urgency_level"] = "moyenne"
        if analysis.get("difficulty_level") not in valid_level:
            analysis["difficulty_level"] = "moyenne"
        if analysis.get("effort_level") not in valid_level:
            analysis["effort_level"] = "moyenne"

        for score_field in ("eligibility_score", "strategic_interest"):
            val = analysis.get(score_field)
            if not isinstance(val, (int, float)):
                analysis[score_field] = 50
            else:
                analysis[score_field] = max(0, min(100, int(val)))

        analysis["model"] = "claude-opus-4-5"

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        logger.error(f"[Analyze] JSON invalide : {e}")
        raise HTTPException(status_code=502, detail="L'IA a retourné une réponse invalide. Réessayez.")
    except Exception as e:
        logger.error(f"[Analyze] Erreur IA : {e}")
        raise HTTPException(status_code=502, detail=f"Erreur lors de l'analyse IA : {str(e)}")

    # Stocker l'analyse dans la base
    from datetime import datetime as datetime_class
    device.decision_analysis = analysis
    device.decision_analyzed_at = datetime_class.utcnow()
    await db.commit()
    await db.refresh(device)

    logger.info(f"[Analyze] '{device.title}' → {analysis.get('go_no_go')} / priorité {analysis.get('recommended_priority')}")
    return device


@router.post("/{device_id}/rewrite", response_model=DeviceResponse)
async def rewrite_device(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin", "editor"])),
):
    """
    Lance la reformulation IA de la fiche (toutes les sections métier).
    Appelle l'AIRewriter et stocke le résultat dans ai_rewritten_sections_json.
    """
    from app.services.ai_rewriter import AIRewriter, REWRITE_DONE, REWRITE_NEEDS_REVIEW
    from datetime import datetime as datetime_class

    service = DeviceService(db)
    device = await service.get_by_id(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Dispositif introuvable")

    rewriter = AIRewriter()
    if not rewriter.can_rewrite():
        raise HTTPException(
            status_code=422,
            detail="Le service de reformulation IA n'est pas configuré (clé API manquante)."
        )

    # Build device dict for the rewriter
    device_dict = {
        "id": str(device.id),
        "title": device.title,
        "organism": device.organism,
        "country": device.country,
        "device_type": device.device_type,
        "status": device.status,
        "close_date": str(device.close_date or ""),
        "amount_min": device.amount_min,
        "amount_max": device.amount_max,
        "currency": device.currency,
        "content_sections_json": device.content_sections_json or [],
    }

    result = await rewriter.rewrite_device(device_dict)

    device.ai_rewritten_sections_json = result.sections
    device.ai_rewrite_status = result.status
    device.ai_rewrite_model = result.model
    device.ai_rewrite_checked_at = result.checked_at

    await db.commit()
    await db.refresh(device)

    status_label = "reformulée" if result.status in (REWRITE_DONE, REWRITE_NEEDS_REVIEW) else "échouée"
    logger.info(f"[Rewrite] '{device.title}' → {result.status} ({len(result.sections)} sections, {status_label})")
    return device


@router.delete("/{device_id}", status_code=204)
async def delete_device(
    device_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(["admin"])),
):
    await DeviceService(db).delete(device_id)
